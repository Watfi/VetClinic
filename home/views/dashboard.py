"""Dashboard and reports — admin-only ORM versions."""

import io
from collections import Counter
from datetime import datetime, timedelta

from django.conf import settings
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from home.models import Cita, HistoriaClinica, Paciente, Peluqueria, Producto, Usuario, Venta, VentaItem

from ._helpers import admin_required, vet_or_admin_required


@vet_or_admin_required
def index(request):
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())

    total_users = Usuario.objects.count()
    total_clientes = Usuario.objects.filter(rol=Usuario.ROL_PELUQUERO).count()
    total_veterinarios = Usuario.objects.filter(rol=Usuario.ROL_VET).count()
    total_admins = Usuario.objects.filter(rol=Usuario.ROL_ADMIN).count()

    total_mascotas = Paciente.objects.count()
    especies_counter = Counter(
        Paciente.objects.values_list("especie", flat=True)
    )
    especies_data = [
        {"especie": k or "Desconocido", "cantidad": v}
        for k, v in especies_counter.most_common(5)
    ]

    citas_qs = Cita.objects.all()
    total_citas = citas_qs.count()
    citas_pendientes = citas_qs.filter(estado="Pendiente").count()
    citas_completadas = citas_qs.filter(estado="Completada").count()
    citas_canceladas = citas_qs.filter(estado="Cancelada").count()
    citas_hoy = citas_qs.filter(fecha__date=today).count()
    citas_semana = citas_qs.filter(fecha__date__gte=week_start).count()

    # Citas de hoy (timeline)
    citas_hoy_list = []
    for c in (Cita.objects
              .filter(fecha__date=today)
              .select_related("paciente", "paciente__owner", "veterinario")
              .order_by("fecha")):
        citas_hoy_list.append({
            "hora": timezone.localtime(c.fecha).strftime("%H:%M"),
            "tipo": (c.tipo or "Cita")[:10],
            "mascota": c.paciente.nombre if c.paciente else "—",
            "propietario": (c.paciente.owner.nombre or c.paciente.owner.user) if (c.paciente and c.paciente.owner) else "—",
            "veterinario": (c.veterinario.nombre or c.veterinario.user) if c.veterinario else "—",
            "estado": c.estado,
        })

    return render(request, "index.html", {
        "rol": request.session.get("rol"),
        "username": request.session.get("user"),
        "total_users": total_users,
        "total_clientes": total_clientes,
        "total_veterinarios_users": total_veterinarios,
        "total_admins": total_admins,
        "total_mascotas": total_mascotas,
        "especies_data": especies_data,
        "total_veterinarios": total_veterinarios,
        "total_citas": total_citas,
        "citas_pendientes": citas_pendientes,
        "citas_completadas": citas_completadas,
        "citas_canceladas": citas_canceladas,
        "citas_hoy": citas_hoy,
        "citas_semana": citas_semana,
        "citas_hoy_list": citas_hoy_list,
    })


REPORT_TYPES = [
    ("appointments", "Citas"),
    ("citas_peluqueria", "Citas y Peluquería por día"),
    ("pets", "Mascotas"),
    ("veterinarians", "Veterinarios"),
    ("users", "Usuarios"),
    ("medical_history", "Historias Clínicas"),
    ("sales", "Ventas"),
    ("inventory", "Inventario"),
]


def _excel_response(rows, headers, sheet_name, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="556B2F", end_color="556B2F", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=14)

    # Add business info at top
    business = getattr(settings, "BUSINESS_NAME", "Kane Agropet")
    business_nit = getattr(settings, "BUSINESS_NIT", "")
    business_phone = getattr(settings, "BUSINESS_PHONE", "")
    business_email = getattr(settings, "BUSINESS_EMAIL", "")

    row_offset = 0
    ws.cell(row=1, column=1, value=business).font = title_font
    row_offset = 1
    if business_nit or business_phone or business_email:
        info_text = " | ".join(filter(None, [
            f"NIT: {business_nit}" if business_nit else "",
            f"Tel: {business_phone}" if business_phone else "",
            f"Email: {business_email}" if business_email else "",
        ]))
        if info_text:
            ws.cell(row=2, column=1, value=info_text)
            row_offset = 2

    # Add headers
    header_row = row_offset + 2
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Add data
    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in rows:
            v = row[col - 1] if col - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@vet_or_admin_required
def reports(request):
    # Default to new tabbed UI; legacy type= still supported for excel backwards-compat.
    report_type = request.GET.get("type", "")
    tab = request.GET.get("tab", "inventory")
    start_date = request.GET.get("start_date", "") or request.GET.get("desde", "")
    end_date = request.GET.get("end_date", "") or request.GET.get("hasta", "")
    status = request.GET.get("status", "")
    veterinarian_id = request.GET.get("veterinarian", "")
    fmt = request.GET.get("format", "")
    # If only `tab` is set without legacy `type`, map tab → type so excel etc. still work
    if not report_type:
        report_type = {
            "inventory": "inventory",
            "sales_products": "sales",
            "citas_income": "appointments",
            "earnings": "earnings",
            "owners": "owners",
        }.get(tab, "inventory")

    context = {
        "rol": request.session.get("rol"),
        "username": request.session.get("user"),
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "status": status,
        "veterinarian_id": veterinarian_id,
        "report_types": REPORT_TYPES,
    }

    excel_rows = None
    excel_headers = None
    excel_sheet = report_type

    if report_type == "appointments":
        qs = Cita.objects.select_related("paciente", "paciente__owner", "veterinario")
        if start_date:
            qs = qs.filter(fecha__date__gte=start_date)
        if end_date:
            qs = qs.filter(fecha__date__lte=end_date)
        if status:
            qs = qs.filter(estado=status)
        if veterinarian_id:
            qs = qs.filter(veterinario_id=veterinarian_id)

        citas_data = []
        excel_rows = []
        for c in qs.order_by("-fecha"):
            row = {
                "id_str": str(c.id),
                "fecha_formatted": timezone.localtime(c.fecha).strftime("%Y-%m-%d"),
                "hora_formatted": timezone.localtime(c.fecha).strftime("%H:%M"),
                "estado": c.estado,
                "motivo": c.motivo,
                "duracion": c.duracion,
                "observacion": c.observacion,
                "mascota_nombre": c.paciente.nombre if c.paciente else "Desconocido",
                "mascota_especie": c.paciente.especie if c.paciente else "No aplica",
                "owner_name": c.paciente.owner.user if c.paciente and c.paciente.owner else "Desconocido",
                "veterinario_nombre": c.veterinario.nombre if c.veterinario else "Desconocido",
                "veterinario_especialidad": c.veterinario.especialidad if c.veterinario else "No aplica",
            }
            citas_data.append(row)
            excel_rows.append([
                row["fecha_formatted"], row["hora_formatted"], row["estado"],
                row["mascota_nombre"], row["mascota_especie"], row["owner_name"],
                row["veterinario_nombre"], row["veterinario_especialidad"],
                row["motivo"], row["duracion"], row["observacion"],
            ])
        excel_headers = ["Fecha", "Hora", "Estado", "Mascota", "Especie", "Propietario",
                         "Veterinario", "Especialidad", "Motivo", "Duración (h)", "Observación"]

        agg = qs.values("estado").annotate(count=Count("id"))
        stats = {row["estado"]: row["count"] for row in agg}
        context.update({
            "citas_data": citas_data,
            "total_citas": len(citas_data),
            "total_pendientes": stats.get("Pendiente", 0),
            "total_completadas": stats.get("Completada", 0),
            "total_canceladas": stats.get("Cancelada", 0),
        })

    elif report_type == "pets":
        mascotas_data = []
        excel_rows = []
        for m in Paciente.objects.select_related("owner").annotate(total_citas=Count("citas")):
            row = {
                "id_str": str(m.id),
                "nombre": m.nombre,
                "especie": m.especie,
                "raza": m.raza,
                "owner_name": m.owner.user if m.owner else "Desconocido",
                "owner_email": m.owner.email if m.owner else "No aplica",
                "total_citas": m.total_citas,
            }
            mascotas_data.append(row)
            excel_rows.append([row["nombre"], row["especie"], row["raza"],
                               row["owner_name"], row["owner_email"], row["total_citas"]])
        excel_headers = ["Nombre", "Especie", "Raza", "Propietario", "Email", "Total Citas"]
        especies_counter = dict(
            Paciente.objects.values_list("especie")
            .annotate(c=Count("id"))
            .values_list("especie", "c")
        )
        context.update({
            "mascotas_data": mascotas_data,
            "total_mascotas": len(mascotas_data),
            "especies_counter": especies_counter,
        })

    elif report_type == "veterinarians":
        vets_data = []
        excel_rows = []
        for v in Usuario.objects.filter(rol=Usuario.ROL_VET):
            citas_qs = Cita.objects.filter(veterinario=v)
            row = {
                "id_str": str(v.id),
                "nombre": v.nombre or v.user,
                "Email": v.email,
                "Phone": v.phone,
                "especialidad": v.especialidad,
                "total_citas": citas_qs.count(),
                "citas_pendientes": citas_qs.filter(estado="Pendiente").count(),
                "citas_completadas": citas_qs.filter(estado="Completada").count(),
            }
            vets_data.append(row)
            excel_rows.append([row["nombre"], row["Email"], row["Phone"], row["especialidad"],
                               row["total_citas"], row["citas_pendientes"], row["citas_completadas"]])
        excel_headers = ["Nombre", "Email", "Teléfono", "Especialidad",
                         "Total Citas", "Pendientes", "Completadas"]
        vets_data.sort(key=lambda x: x["total_citas"], reverse=True)
        context.update({
            "vets_data": vets_data,
            "total_veterinarios": len(vets_data),
            "total_citas_all": sum(v["total_citas"] for v in vets_data),
        })

    elif report_type == "users":
        users_data = []
        excel_rows = []
        for u in Usuario.objects.all():
            mascotas_count = Paciente.objects.filter(owner=u).count()
            row = {
                "id_str": str(u.id),
                "User": u.user,
                "Email": u.email,
                "Rol": u.rol,
                "Phone": u.phone,
                "total_mascotas": mascotas_count,
            }
            users_data.append(row)
            excel_rows.append([row["User"], row["Email"], row["Rol"], row["Phone"], row["total_mascotas"]])
        excel_headers = ["Usuario", "Email", "Rol", "Teléfono", "Total Mascotas"]
        context.update({
            "users_data": users_data,
            "total_users": len(users_data),
        })

    elif report_type == "medical_history":
        historias_data = []
        excel_rows = []
        for h in HistoriaClinica.objects.select_related("paciente"):
            row = {
                "id_str": str(h.id),
                "hc_numero": h.hc_numero,
                "fecha": h.fecha.strftime("%Y-%m-%d") if h.fecha else "",
                "paciente_nombre": h.paciente.nombre if h.paciente else "",
                "paciente_especie": h.paciente.especie if h.paciente else "",
                "diagnostico": h.diagnostico,
            }
            historias_data.append(row)
            excel_rows.append([row["hc_numero"], row["fecha"], row["paciente_nombre"],
                               row["paciente_especie"], row["diagnostico"]])
        excel_headers = ["HC #", "Fecha", "Paciente", "Especie", "Diagnóstico"]
        context.update({
            "historias_data": historias_data,
            "total_historias": len(historias_data),
        })

    elif report_type == "sales":
        qs = Venta.objects.all()
        if start_date:
            qs = qs.filter(fecha__date__gte=start_date)
        if end_date:
            qs = qs.filter(fecha__date__lte=end_date)
        ventas_data = []
        excel_rows = []
        total_costo = 0.0
        total_venta = 0.0
        for v in qs.order_by("-fecha").prefetch_related("items__producto"):
            costo = 0.0
            venta_val = 0.0
            for it in v.items.all():
                pc = float(it.producto.precio_compra) if it.producto else 0.0
                costo += pc * it.cantidad
                venta_val += float(it.subtotal)
            margen = venta_val - costo
            if v.estado == "Pagada":
                total_costo += costo
                total_venta += venta_val
            ventas_data.append({
                "numero": v.numero,
                "fecha": timezone.localtime(v.fecha).strftime("%Y-%m-%d %H:%M"),
                "cliente": v.cliente_nombre,
                "estado": v.estado,
                "total": float(v.total),
                "valor_compra": costo,
                "valor_venta": venta_val,
                "margen": margen,
            })
            excel_rows.append([
                v.numero, timezone.localtime(v.fecha).strftime("%Y-%m-%d %H:%M"),
                v.cliente_nombre, v.cliente_email, v.metodo_pago, v.estado,
                float(v.subtotal), float(v.impuestos), float(v.total),
                costo, venta_val, margen,
            ])
        excel_headers = ["Número", "Fecha", "Cliente", "Email", "Método Pago",
                         "Estado", "Subtotal", "Impuestos", "Total",
                         "Valor Compra", "Valor Venta", "Margen"]
        agg_total = qs.filter(estado="Pagada").aggregate(total=Sum("total"))["total"] or 0
        context.update({
            "ventas_data": ventas_data,
            "total_ventas": qs.count(),
            "total_recaudado": float(agg_total),
            "total_costo": total_costo,
            "total_valor_venta": total_venta,
            "total_margen": total_venta - total_costo,
        })

    elif report_type == "inventory":
        productos_data = []
        excel_rows = []
        for p in Producto.objects.select_related("categoria"):
            margen_unit = float(p.precio) - float(p.precio_compra)
            productos_data.append({
                "sku": p.sku,
                "nombre": p.nombre,
                "categoria": p.categoria.nombre if p.categoria else "",
                "stock": p.stock,
                "precio_compra": float(p.precio_compra),
                "precio": float(p.precio),
                "margen": margen_unit,
                "activo": p.activo,
            })
            excel_rows.append([
                p.sku, p.nombre, p.categoria.nombre if p.categoria else "",
                p.stock, p.stock_minimo,
                float(p.precio_compra), float(p.precio), margen_unit,
                "Sí" if p.activo else "No",
            ])
        excel_headers = ["SKU", "Nombre", "Categoría", "Stock", "Stock Mínimo",
                         "Valor Compra", "Valor Venta", "Margen", "Activo"]
        context.update({
            "productos_data": productos_data,
            "total_productos": len(productos_data),
        })

    elif report_type == "citas_peluqueria":
        from collections import defaultdict
        from datetime import date as _date

        qs_citas = Cita.objects.all()
        qs_pel = Peluqueria.objects.all()
        if start_date:
            qs_citas = qs_citas.filter(fecha__date__gte=start_date)
            qs_pel = qs_pel.filter(fecha__gte=start_date)
        if end_date:
            qs_citas = qs_citas.filter(fecha__date__lte=end_date)
            qs_pel = qs_pel.filter(fecha__lte=end_date)

        # Citas por día y tipo
        citas_por_dia = defaultdict(lambda: {"Medica": 0, "Peluqueria": 0})
        tipo_counter = Counter()
        motivo_counter = Counter()
        for c in qs_citas:
            fecha_local = timezone.localtime(c.fecha)
            dia = fecha_local.date().isoformat()
            t = c.tipo or "Medica"
            citas_por_dia[dia][t] += 1
            tipo_counter[t] += 1
            if c.motivo:
                motivo_counter[c.motivo] += 1

        # Peluquerías por día
        pel_por_dia = defaultdict(int)
        servicio_counter = Counter()
        for p in qs_pel:
            pel_por_dia[p.fecha.isoformat()] += 1
            servicio_counter[p.servicio] += 1

        # Unificar días
        all_days = sorted(set(list(citas_por_dia.keys()) + list(pel_por_dia.keys())))

        chart_labels = all_days
        chart_medicas = [citas_por_dia[d]["Medica"] for d in all_days]
        chart_peluqueria_cita = [citas_por_dia[d]["Peluqueria"] for d in all_days]
        chart_peluqueria_spa = [pel_por_dia[d] for d in all_days]

        # Tabla detallada
        detalle = []
        for d in all_days:
            detalle.append({
                "fecha": d,
                "medicas": citas_por_dia[d]["Medica"],
                "peluqueria_cita": citas_por_dia[d]["Peluqueria"],
                "peluqueria_spa": pel_por_dia[d],
                "total": citas_por_dia[d]["Medica"] + citas_por_dia[d]["Peluqueria"] + pel_por_dia[d],
            })

        # Tarifas para costo estimado por tipo
        from home.models import TarifaCita
        tarifas_dict = {t.tipo: float(t.precio) for t in TarifaCita.objects.filter(activo=True)}

        excel_rows = [[d["fecha"], d["medicas"], d["peluqueria_cita"], d["peluqueria_spa"], d["total"]] for d in detalle]
        excel_headers = ["Fecha", "Citas médicas", "Citas peluquería", "Peluquería spa", "Total"]

        context.update({
            "detalle": detalle,
            "chart_labels": chart_labels,
            "chart_medicas": chart_medicas,
            "chart_peluqueria_cita": chart_peluqueria_cita,
            "chart_peluqueria_spa": chart_peluqueria_spa,
            "tipo_counter": dict(tipo_counter),
            "motivo_counter": dict(motivo_counter.most_common(10)),
            "servicio_counter": dict(servicio_counter),
            "tarifas_dict": tarifas_dict,
            "total_medicas": sum(chart_medicas),
            "total_pel_cita": sum(chart_peluqueria_cita),
            "total_pel_spa": sum(chart_peluqueria_spa),
        })

    # ── Aggregations for the new tabbed UI ─────────────────────────────────
    from collections import defaultdict
    from home.models import Propietario, TarifaCita
    import json as _json

    # All active tarifas (type → price)
    _tarifas_qs = TarifaCita.objects.filter(activo=True)
    _tarifas = {t.tipo: float(t.precio) for t in _tarifas_qs}

    def _cita_tarifa_key(c):
        """Map a cita to the best matching tarifa key.
        Priority: motivo_consulta on related HC > motivo > tipo.
        """
        # Try matching motivo against tarifa types
        motivo = (c.motivo or "").strip()
        if motivo in _tarifas:
            return motivo
        # Fallback to cita tipo mapping
        tipo = c.tipo or "Medica"
        if tipo == "Medica":
            return "Consulta general"
        elif tipo == "Peluqueria":
            return "Peluquería básica"
        return motivo or tipo

    def _cita_income(c):
        """Resolve cita income: explicit price > tarifa for matched key > 0."""
        amt = float(getattr(c, "precio", 0) or 0)
        if amt > 0:
            return amt
        key = _cita_tarifa_key(c)
        return _tarifas.get(key, 0.0)

    # Date filter helpers
    def _filter_dt(qs, field):
        if start_date:
            qs = qs.filter(**{f"{field}__gte": start_date})
        if end_date:
            qs = qs.filter(**{f"{field}__lte": end_date})
        return qs

    # Top productos vendidos (bar chart)
    items_qs = VentaItem.objects.select_related("producto", "venta").filter(venta__estado="Pagada")
    if start_date: items_qs = items_qs.filter(venta__fecha__date__gte=start_date)
    if end_date:   items_qs = items_qs.filter(venta__fecha__date__lte=end_date)
    prod_agg = defaultdict(lambda: {"qty": 0, "ingresos": 0.0, "costo": 0.0})
    for it in items_qs:
        key = it.nombre or (it.producto.nombre if it.producto else "—")
        prod_agg[key]["qty"] += it.cantidad
        prod_agg[key]["ingresos"] += float(it.subtotal)
        pc = float(it.producto.precio_compra) if it.producto else 0.0
        prod_agg[key]["costo"] += pc * it.cantidad
    top_productos = sorted(
        ({"nombre": k, **v, "margen": v["ingresos"] - v["costo"]} for k, v in prod_agg.items()),
        key=lambda x: x["ingresos"], reverse=True,
    )[:10]

    # Ingresos por citas — por mes, semana, año y por tipo de tarifa
    citas_q = Cita.objects.exclude(estado="Cancelada")
    if start_date: citas_q = citas_q.filter(fecha__date__gte=start_date)
    if end_date:   citas_q = citas_q.filter(fecha__date__lte=end_date)
    citas_por_mes_dict = defaultdict(float)
    citas_por_semana_dict = defaultdict(float)
    citas_por_anio_dict = defaultdict(float)
    citas_por_tipo = defaultdict(lambda: {"count": 0, "ingresos": 0.0, "precio_tarifa": 0.0})

    # Initialize with ALL active tarifas so they all appear even if count=0
    for tarifa_tipo, tarifa_precio in _tarifas.items():
        citas_por_tipo[tarifa_tipo]["precio_tarifa"] = tarifa_precio

    for c in citas_q:
        fecha_local = timezone.localtime(c.fecha)
        ym = fecha_local.strftime("%Y-%m")
        yw = fecha_local.strftime("%Y-W%W")
        yyyy = fecha_local.strftime("%Y")
        income = _cita_income(c)
        citas_por_mes_dict[ym] += income
        citas_por_semana_dict[yw] += income
        citas_por_anio_dict[yyyy] += income
        t = _cita_tarifa_key(c)
        citas_por_tipo[t]["count"] += 1
        citas_por_tipo[t]["ingresos"] += income
        if t in _tarifas:
            citas_por_tipo[t]["precio_tarifa"] = _tarifas[t]

    citas_meses_sorted = sorted(citas_por_mes_dict.keys())
    citas_semanas_sorted = sorted(citas_por_semana_dict.keys())
    citas_anios_sorted = sorted(citas_por_anio_dict.keys())
    citas_chart = {
        "labels": citas_meses_sorted,
        "values": [citas_por_mes_dict[m] for m in citas_meses_sorted],
    }
    citas_tipo_table = sorted(
        ({"tipo": k, **v} for k, v in citas_por_tipo.items()),
        key=lambda x: x["ingresos"], reverse=True,
    )
    total_citas_count = sum(r["count"] for r in citas_tipo_table)

    # Ganancias totales
    ventas_q = Venta.objects.filter(estado="Pagada")
    if start_date: ventas_q = ventas_q.filter(fecha__date__gte=start_date)
    if end_date:   ventas_q = ventas_q.filter(fecha__date__lte=end_date)
    ingresos_ventas = float(ventas_q.aggregate(total=Sum("total"))["total"] or 0)
    costo_productos = sum(p["costo"] for p in prod_agg.values())
    ingresos_citas = sum(citas_por_mes_dict.values())
    utilidad_bruta = (ingresos_ventas - costo_productos) + ingresos_citas

    # Propietarios — totales y altas por mes
    prop_qs = Propietario.objects.all()
    total_propietarios = prop_qs.count()
    prop_por_mes = defaultdict(int)
    if hasattr(Propietario, "creado_en") or hasattr(Propietario, "fecha_creacion"):
        date_field = "creado_en" if hasattr(Propietario, "creado_en") else "fecha_creacion"
        for p in prop_qs.values_list(date_field, flat=True):
            if p:
                prop_por_mes[p.strftime("%Y-%m")] += 1
    prop_meses = sorted(prop_por_mes.keys())
    prop_chart = {
        "labels": prop_meses,
        "values": [prop_por_mes[m] for m in prop_meses],
    }
    # Top propietarios por gasto
    top_owners = []
    owner_spend = defaultdict(float)
    owner_name = {}
    for v in ventas_q.select_related("cliente"):
        key = v.cliente_id or v.cliente_nombre or "—"
        owner_spend[key] += float(v.total)
        owner_name[key] = v.cliente_nombre or (v.cliente.user if v.cliente else "—")
    for k, total in sorted(owner_spend.items(), key=lambda x: x[1], reverse=True)[:10]:
        top_owners.append({"nombre": owner_name.get(k, "—"), "total": total})

    # Inventory snapshot for Inventario tab (also reused above when type=inventory)
    inv_total_value = sum((float(p.precio) * p.stock) for p in Producto.objects.all())
    inv_total_cost  = sum((float(p.precio_compra) * p.stock) for p in Producto.objects.all())
    inv_low_stock   = sum(1 for p in Producto.objects.all() if p.stock <= (p.stock_minimo or 0))

    context.update({
        "tab": tab,
        "start_date": start_date,
        "end_date": end_date,
        "tarifas_dict_full": _tarifas,
        # Sales by product
        "top_productos": top_productos,
        "top_productos_json": _json.dumps({
            "labels": [p["nombre"] for p in top_productos],
            "ingresos": [p["ingresos"] for p in top_productos],
            "costo": [p["costo"] for p in top_productos],
        }),
        # Citas income
        "citas_tipo_table": citas_tipo_table,
        "citas_chart_json": _json.dumps(citas_chart),
        "total_ingresos_citas": ingresos_citas,
        "total_citas_count": total_citas_count,
        # Period breakdowns for reports
        "citas_por_mes": citas_por_mes_dict,
        "citas_meses_sorted": citas_meses_sorted,
        "citas_por_semana": citas_por_semana_dict,
        "citas_semanas_sorted": citas_semanas_sorted,
        "citas_por_anio": citas_por_anio_dict,
        "citas_anios_sorted": citas_anios_sorted,
        # Earnings
        "ingresos_ventas": ingresos_ventas,
        "ingresos_citas": ingresos_citas,
        "costo_productos": costo_productos,
        "utilidad_bruta": utilidad_bruta,
        "earnings_chart_json": _json.dumps({
            "labels": ["Ingresos ventas", "Ingresos citas", "Costo productos"],
            "values": [ingresos_ventas, ingresos_citas, costo_productos],
        }),
        # Propietarios
        "total_propietarios": total_propietarios,
        "prop_chart_json": _json.dumps(prop_chart),
        "top_owners": top_owners,
        # Inventory KPIs
        "inv_total_value": inv_total_value,
        "inv_total_cost": inv_total_cost,
        "inv_low_stock": inv_low_stock,
    })

    # ── PDF export for tabs (except inventory which stays excel) ──────────
    if fmt == "pdf" and tab in ("sales_products", "citas_income", "earnings", "owners"):
        return _report_pdf(tab, context)

    if fmt == "excel" and excel_rows is not None:
        filename = f"reporte_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return _excel_response(excel_rows, excel_headers, excel_sheet, filename)

    veterinarios_list = list(
        Usuario.objects.filter(rol=Usuario.ROL_VET)
        .values("id", "nombre", "user", "especialidad")
    )
    for v in veterinarios_list:
        v["id_str"] = str(v["id"])
        v["nombre"] = v["nombre"] or v["user"]
    context["veterinarios_list"] = veterinarios_list

    return render(request, "reports.html", context)


def _report_pdf(tab, ctx):
    """Generate a PDF report for the given tab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(LETTER),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    business = getattr(settings, "BUSINESS_NAME", "Kane Agropet")
    business_nit = getattr(settings, "BUSINESS_NIT", "")
    business_phone = getattr(settings, "BUSINESS_PHONE", "")
    business_email = getattr(settings, "BUSINESS_EMAIL", "")

    story.append(Paragraph(f"<b>{business}</b>", styles["Title"]))

    # Business info
    info_lines = []
    if business_nit:
        info_lines.append(f"NIT: {business_nit}")
    if business_phone:
        info_lines.append(f"Tel: {business_phone}")
    if business_email:
        info_lines.append(f"Email: {business_email}")
    if info_lines:
        story.append(Paragraph(" | ".join(info_lines), styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    date_range = ""
    if ctx.get("start_date") or ctx.get("end_date"):
        ds = ctx.get("start_date", "—")
        de = ctx.get("end_date", "—")
        date_range = f"  |  Período: {ds} a {de}"

    olive_color = colors.HexColor("#556B2F")
    header_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), olive_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
    ])

    if tab == "sales_products":
        story.append(Paragraph(f"<b>Reporte: Ventas de productos</b>{date_range}", styles["Heading2"]))
        story.append(Spacer(1, 0.4*cm))
        data = [["Producto", "Cantidad", "Ingresos", "Costo", "Margen"]]
        for p in ctx.get("top_productos", []):
            data.append([
                p["nombre"], str(p["qty"]),
                f"${p['ingresos']:,.2f}", f"${p['costo']:,.2f}",
                f"${p['margen']:,.2f}",
            ])
        if len(data) > 1:
            tbl = Table(data, colWidths=[10*cm, 2*cm, 3.5*cm, 3.5*cm, 3.5*cm])
            tbl.setStyle(header_style)
            story.append(tbl)
        else:
            story.append(Paragraph("Sin datos de ventas en el rango.", styles["Normal"]))

    elif tab == "citas_income":
        story.append(Paragraph(f"<b>Reporte: Ingresos por citas</b>{date_range}", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"Total ingresos: <b>${ctx.get('total_ingresos_citas', 0):,.2f}</b>", styles["Normal"]))
        story.append(Spacer(1, 0.4*cm))
        data = [["Tipo de cita", "Precio tarifa", "Cantidad", "Ingresos"]]
        for r in ctx.get("citas_tipo_table", []):
            data.append([
                r["tipo"], f"${r['precio_tarifa']:,.2f}",
                str(r["count"]), f"${r['ingresos']:,.2f}",
            ])
        total_count = sum(r["count"] for r in ctx.get("citas_tipo_table", []))
        total_income = ctx.get("total_ingresos_citas", 0)
        data.append(["TOTAL", "", str(total_count), f"${total_income:,.2f}"])
        tbl = Table(data, colWidths=[8*cm, 4*cm, 3*cm, 4*cm])
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), olive_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8e8e8")),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
        ])
        tbl.setStyle(style)
        story.append(tbl)

    elif tab == "earnings":
        story.append(Paragraph(f"<b>Reporte: Ganancias totales</b>{date_range}", styles["Heading2"]))
        story.append(Spacer(1, 0.4*cm))

        # Resumen general
        data = [
            ["Concepto", "Monto"],
            ["Ingresos por ventas", f"${ctx.get('ingresos_ventas', 0):,.2f}"],
            ["Ingresos por citas", f"${ctx.get('ingresos_citas', 0):,.2f}"],
            ["Costo de productos vendidos", f"- ${ctx.get('costo_productos', 0):,.2f}"],
            ["UTILIDAD BRUTA", f"${ctx.get('utilidad_bruta', 0):,.2f}"],
        ]
        tbl = Table(data, colWidths=[12*cm, 6*cm])
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), olive_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, -1), (-1, -1), 11),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8e8e8")),
            ("FONTSIZE", (0, 1), (-1, -2), 9),
        ])
        tbl.setStyle(style)
        story.append(tbl)

        # Breakdown by period
        story.append(Spacer(1, 0.6*cm))
        story.append(Paragraph("<b>Desglose por período</b>", styles["Heading3"]))
        story.append(Spacer(1, 0.2*cm))

        import json

        # Monthly breakdown
        citas_chart = ctx.get("citas_chart_json")
        if citas_chart:
            try:
                chart_data = json.loads(citas_chart)
                months_data = list(zip(chart_data.get("labels", []), chart_data.get("values", [])))
                if months_data:
                    story.append(Paragraph("<b>Ingresos por mes:</b>", styles["Normal"]))
                    month_table_data = [["Mes", "Ingresos por citas"]]
                    for month, value in months_data:
                        month_table_data.append([month, f"${value:,.2f}"])
                    month_tbl = Table(month_table_data, colWidths=[6*cm, 6*cm])
                    month_style = TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#999999")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ])
                    month_tbl.setStyle(month_style)
                    story.append(month_tbl)
                    story.append(Spacer(1, 0.3*cm))
            except:
                pass

        # Weekly breakdown
        citas_semanas = ctx.get("citas_semanas_sorted", [])
        citas_por_semana = ctx.get("citas_por_semana", {})
        if citas_semanas:
            story.append(Paragraph("<b>Ingresos por semana:</b>", styles["Normal"]))
            week_table_data = [["Semana", "Ingresos por citas"]]
            for week in citas_semanas:
                week_table_data.append([week, f"${citas_por_semana[week]:,.2f}"])
            week_tbl = Table(week_table_data, colWidths=[6*cm, 6*cm])
            week_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#999999")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ])
            week_tbl.setStyle(week_style)
            story.append(week_tbl)
            story.append(Spacer(1, 0.3*cm))

        # Yearly breakdown
        citas_anios = ctx.get("citas_anios_sorted", [])
        citas_por_anio = ctx.get("citas_por_anio", {})
        if citas_anios:
            story.append(Paragraph("<b>Ingresos por año:</b>", styles["Normal"]))
            year_table_data = [["Año", "Ingresos por citas"]]
            for year in citas_anios:
                year_table_data.append([year, f"${citas_por_anio[year]:,.2f}"])
            year_tbl = Table(year_table_data, colWidths=[6*cm, 6*cm])
            year_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#999999")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ])
            year_tbl.setStyle(year_style)
            story.append(year_tbl)
            story.append(Spacer(1, 0.3*cm))

    elif tab == "owners":
        story.append(Paragraph(f"<b>Reporte: Propietarios</b>{date_range}", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"Total propietarios registrados: <b>{ctx.get('total_propietarios', 0)}</b>", styles["Normal"]))
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("<b>Top 10 por gasto</b>", styles["Heading3"]))
        data = [["#", "Cliente", "Total gastado"]]
        for i, o in enumerate(ctx.get("top_owners", []), 1):
            data.append([str(i), o["nombre"], f"${o['total']:,.2f}"])
        if len(data) > 1:
            tbl = Table(data, colWidths=[1.5*cm, 12*cm, 5*cm])
            tbl.setStyle(header_style)
            story.append(tbl)
        else:
            story.append(Paragraph("Sin datos de propietarios.", styles["Normal"]))

    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))

    doc.build(story)
    tab_names = {
        "sales_products": "ventas_productos",
        "citas_income": "ingresos_citas",
        "earnings": "ganancias",
        "owners": "propietarios",
    }
    filename = f"reporte_{tab_names.get(tab, tab)}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp

